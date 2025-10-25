import os
from dotenv import load_dotenv
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Carica variabili d'ambiente
load_dotenv()
TOKEN = os.getenv("TELEGRAM_TOKEN")

# Nome del file principale che salva tutti i dati
FILE_EXCEL = "estratto_conto.xlsx"

# Carica o crea il file Excel
if os.path.exists(FILE_EXCEL):
    wb = load_workbook(FILE_EXCEL)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["user_id", "username", "movimento", "data_ora"])  # intestazioni
    wb.save(FILE_EXCEL)

# Funzioni di utilità
def salva_movimento(user_id: int, username: str, valore: float):
    ora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([user_id, username, valore, ora])
    wb.save(FILE_EXCEL)

def leggi_movimenti_completo():
    movimenti = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        movimenti.append({"valore": row[2], "username": row[1], "data_ora": row[3]})
    return movimenti

def estratto_conto_completo():
    movimenti = leggi_movimenti_completo()
    totale_entrate = sum([m["valore"] for m in movimenti if m["valore"] > 0])
    totale_uscite = sum([m["valore"] for m in movimenti if m["valore"] < 0])
    saldo = totale_entrate + totale_uscite
    return movimenti, totale_entrate, totale_uscite, saldo

def annulla_ultimo(user_id: int):
    rows = list(ws.iter_rows(min_row=2))
    for row in reversed(rows):
        if row[0].value == user_id:
            ws.delete_rows(row[0].row, 1)
            wb.save(FILE_EXCEL)
            return True
    return False

def reset_tutto():
    global ws
    wb.remove(ws)
    ws = wb.create_sheet("Sheet1")
    ws.append(["user_id", "username", "movimento", "data_ora"])
    wb.save(FILE_EXCEL)

def crea_file_excel_completo():
    movimenti, totale_entrate, totale_uscite, saldo = estratto_conto_completo()
    
    wb_user = Workbook()
    ws_user = wb_user.active
    ws_user.title = "Estratto Conto Completo"
    
    ws_user.append(["Tipo", "Importo", "Utente", "Data/Ora"])
    
    for m in movimenti:
        tipo = "Entrata" if m["valore"] > 0 else "Uscita"
        ws_user.append([tipo, m["valore"], m["username"], m["data_ora"]])
    
    ws_user.append([])
    ws_user.append(["Totale Entrate", totale_entrate])
    ws_user.append(["Totale Uscite", totale_uscite])
    ws_user.append(["Saldo Finale", saldo])
    
    filename = "estratto_conto_completo.xlsx"
    wb_user.save(filename)
    return filename

# Comandi del bot
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Ciao! Sono TotalX Estratto Conto Bot Avanzato.\n"
        "Comandi disponibili:\n"
        "/add numero - aggiunge un'entrata\n"
        "/subtract numero - aggiunge un'uscita\n"
        "/total - mostra il saldo totale\n"
        "/report - mostra l'estratto conto completo\n"
        "/export - ricevi un file Excel con l'estratto conto completo\n"
        "/undo - annulla l'ultima operazione\n"
        "/reset - azzera tutto e crea un nuovo foglio"
    )

async def add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        raw_value = context.args[0].replace(",", ".")
        value = float(raw_value)
        user_id = update.message.from_user.id
        username = update.message.from_user.username or update.message.from_user.first_name
        salva_movimento(user_id, username, value)
        _, totale_entrate, totale_uscite, saldo = estratto_conto_completo()
        await update.message.reply_text(f"Entrata registrata: +{value}\nSaldo totale: {saldo}")
    except (IndexError, ValueError):
        await update.message.reply_text("Errore! Usa /add numero, esempio /add 100 o /add 0,05")

async def subtract(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        raw_value = context.args[0].replace(",", ".")
        value = float(raw_value)
        user_id = update.message.from_user.id
        username = update.message.from_user.username or update.message.from_user.first_name
        salva_movimento(user_id, username, -value)
        _, totale_entrate, totale_uscite, saldo = estratto_conto_completo()
        await update.message.reply_text(f"Uscita registrata: -{value}\nSaldo totale: {saldo}")
    except (IndexError, ValueError):
        await update.message.reply_text("Errore! Usa /subtract numero, esempio /subtract 50 o /subtract 0,05")

async def total(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _, totale_entrate, totale_uscite, saldo = estratto_conto_completo()
    await update.message.reply_text(f"Saldo totale: {saldo}")

async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    movimenti, totale_entrate, totale_uscite, saldo = estratto_conto_completo()
    
    if not movimenti:
        await update.message.reply_text("Nessun movimento registrato.")
        return
    
    report_text = "📄 Estratto Conto Completo\n\n"
    for m in movimenti:
        tipo = "Entrata" if m["valore"] > 0 else "Uscita"
        report_text += f"{tipo}: {m['valore']} ({m['username']} {m['data_ora']})\n"
    
    report_text += f"\nTotale Entrate: {totale_entrate}\nTotale Uscite: {totale_uscite}\nSaldo Totale: {saldo}"
    
    await update.message.reply_text(report_text)

async def export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    filename = crea_file_excel_completo()
    with open(filename, "rb") as file:
        await update.message.reply_document(file, filename=filename)

async def undo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    success = annulla_ultimo(user_id)
    if success:
        await update.message.reply_text("Ultima operazione annullata.")
    else:
        await update.message.reply_text("Nessuna operazione da annullare.")

async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reset_tutto()
    await update.message.reply_text("Tutto azzerato. Nuovo foglio creato.")

# Funzione principale
def main():
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("add", add))
    app.add_handler(CommandHandler("subtract", subtract))
    app.add_handler(CommandHandler("total", total))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(CommandHandler("export", export))
    app.add_handler(CommandHandler("undo", undo))
    app.add_handler(CommandHandler("reset", reset))

    print("Bot avviato...")
    app.run_polling()

if __name__ == "__main__":
    main()
